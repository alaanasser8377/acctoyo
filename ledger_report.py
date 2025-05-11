from flask import Blueprint, render_template, request, redirect, url_for, session, send_file
from database import get_db, ACCOUNT_TYPES
from datetime import datetime
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ledger_bp = Blueprint('ledger', __name__)

def get_active_financial_year_dates():
    """الحصول على تواريخ بداية ونهاية السنة المالية النشطة"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT start_date, end_date
        FROM financial_years
        WHERE is_active = 1
        LIMIT 1
    """)
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return result['start_date'], result['end_date']
    return None, None

def get_parent_accounts(conn):
    """الحصول على قائمة الحسابات الرئيسية (التي لها حسابات فرعية)"""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT a.id, a.code, a.name, a.account_type
        FROM accounts a
        WHERE EXISTS (
            SELECT 1 FROM accounts sub WHERE sub.parent_id = a.id
        )
        AND a.is_active = 1
        ORDER BY a.code
    """)
    return cursor.fetchall()

def get_ledger_data(conn, filters):
    """
    استخراج بيانات دفتر الأستاذ للحساب الرئيسي وحساباته الفرعية
    
    Parameters:
    - conn: اتصال قاعدة البيانات
    - filters: معايير التصفية (parent_account_id, start_date, end_date)
    
    Returns:
    - قاموس يحتوي على بيانات دفتر الأستاذ
    """
    cursor = conn.cursor()
    
    # التحقق من وجود حساب رئيسي محدد
    if not filters.get('parent_account_id'):
        return {
            'parent_account': None,
            'accounts': [],
            'totals': {
                'opening_debit': 0,
                'opening_credit': 0,
                'period_debit': 0,
                'period_credit': 0,
                'total_debit': 0,
                'total_credit': 0,
                'balance_debit': 0,
                'balance_credit': 0
            }
        }
    
    # الحصول على معلومات الحساب الرئيسي
    cursor.execute("""
        SELECT id, code, name, account_type
        FROM accounts
        WHERE id = ?
    """, (filters['parent_account_id'],))
    
    parent_account = cursor.fetchone()
    
    if not parent_account:
        return {
            'parent_account': None,
            'accounts': [],
            'totals': {
                'opening_debit': 0,
                'opening_credit': 0,
                'period_debit': 0,
                'period_credit': 0,
                'total_debit': 0,
                'total_credit': 0,
                'balance_debit': 0,
                'balance_credit': 0
            }
        }
    
    # استخراج الحسابات الفرعية للحساب الرئيسي
    cursor.execute("""
        SELECT id, code, name, account_type, parent_id
        FROM accounts
        WHERE parent_id = ? AND is_active = 1
        ORDER BY code
    """, (parent_account['id'],))
    
    sub_accounts = cursor.fetchall()
    
    # تحضير النتائج
    result_data = {
        'parent_account': parent_account,
        'accounts': [],
        'totals': {
            'opening_debit': 0,
            'opening_credit': 0,
            'period_debit': 0,
            'period_credit': 0,
            'total_debit': 0,
            'total_credit': 0,
            'balance_debit': 0,
            'balance_credit': 0
        }
    }
    
    # معالجة كل حساب فرعي
    for account in sub_accounts:
        account_id = account['id']
        account_type = account['account_type']
        
        # حساب الرصيد الافتتاحي (قبل تاريخ البداية)
        opening_debit = 0
        opening_credit = 0
        
        if filters.get('start_date'):
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN jvd.debit > 0 THEN jvd.debit ELSE 0 END) as total_debit,
                    SUM(CASE WHEN jvd.credit > 0 THEN jvd.credit ELSE 0 END) as total_credit
                FROM 
                    journal_voucher_details jvd
                JOIN 
                    journal_vouchers jv ON jvd.voucher_id = jv.id
                WHERE 
                    jvd.account_id = ? AND jv.voucher_date < ?
            """, (account_id, filters['start_date']))
            
            opening_result = cursor.fetchone()
            if opening_result:
                total_debit = opening_result['total_debit'] if opening_result['total_debit'] is not None else 0
                total_credit = opening_result['total_credit'] if opening_result['total_credit'] is not None else 0
                
                # تحديد الرصيد الافتتاحي بناءً على نوع الحساب
                if account_type in [1, 5]:  # أصول ومصروفات (مدين بطبيعته)
                    if total_debit >= total_credit:
                        opening_debit = total_debit - total_credit
                    else:
                        opening_credit = total_credit - total_debit
                else:  # خصوم وحقوق ملكية وإيرادات (دائن بطبيعته)
                    if total_credit >= total_debit:
                        opening_credit = total_credit - total_debit
                    else:
                        opening_debit = total_debit - total_credit
        
        # حساب حركة الفترة
        period_debit = 0
        period_credit = 0
        
        query_params = [account_id]
        period_query = """
            SELECT 
                SUM(CASE WHEN jvd.debit > 0 THEN jvd.debit ELSE 0 END) as period_debit,
                SUM(CASE WHEN jvd.credit > 0 THEN jvd.credit ELSE 0 END) as period_credit
            FROM 
                journal_voucher_details jvd
            JOIN 
                journal_vouchers jv ON jvd.voucher_id = jv.id
            WHERE 
                jvd.account_id = ?
        """
        
        if filters.get('start_date'):
            period_query += " AND jv.voucher_date >= ?"
            query_params.append(filters['start_date'])
        
        if filters.get('end_date'):
            period_query += " AND jv.voucher_date <= ?"
            query_params.append(filters['end_date'])
        
        cursor.execute(period_query, query_params)
        period_result = cursor.fetchone()
        
        if period_result:
            period_debit = period_result['period_debit'] if period_result['period_debit'] is not None else 0
            period_credit = period_result['period_credit'] if period_result['period_credit'] is not None else 0
        
        # حساب المجاميع والأرصدة
        total_debit = opening_debit + period_debit
        total_credit = opening_credit + period_credit
        
        # تحديد الرصيد النهائي بناءً على نوع الحساب
        balance_debit = 0
        balance_credit = 0
        
        if account_type in [1, 5]:  # أصول ومصروفات (مدين بطبيعته)
            if total_debit >= total_credit:
                balance_debit = total_debit - total_credit
            else:
                balance_credit = total_credit - total_debit
        else:  # خصوم وحقوق ملكية وإيرادات (دائن بطبيعته)
            if total_credit >= total_debit:
                balance_credit = total_credit - total_debit
            else:
                balance_debit = total_debit - total_credit
        
        # إضافة بيانات الحساب إلى النتائج
        account_data = {
            'id': account_id,
            'code': account['code'],
            'name': account['name'],
            'account_type': account_type,
            'account_type_name': ACCOUNT_TYPES.get(account_type, ''),
            'opening_debit': opening_debit,
            'opening_credit': opening_credit,
            'period_debit': period_debit,
            'period_credit': period_credit,
            'total_debit': total_debit,
            'total_credit': total_credit,
            'balance_debit': balance_debit,
            'balance_credit': balance_credit
        }
        
        result_data['accounts'].append(account_data)
        
        # تحديث المجاميع الكلية
        result_data['totals']['opening_debit'] += opening_debit
        result_data['totals']['opening_credit'] += opening_credit
        result_data['totals']['period_debit'] += period_debit
        result_data['totals']['period_credit'] += period_credit
        result_data['totals']['total_debit'] += total_debit
        result_data['totals']['total_credit'] += total_credit
        result_data['totals']['balance_debit'] += balance_debit
        result_data['totals']['balance_credit'] += balance_credit
    
    return result_data

@ledger_bp.route('/reports/ledger', methods=['GET'])
def ledger_report_view():
    """عرض صفحة تقرير دفتر الأستاذ مع إمكانية التصفية"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # إضافة معيار تصفية الحساب الرئيسي إذا تم تحديده
    if request.args.get('parent_account_id'):
        filters['parent_account_id'] = request.args.get('parent_account_id')
    
    # جلب قائمة الحسابات الرئيسية للاختيار
    conn = get_db()
    parent_accounts = get_parent_accounts(conn)
    
    # جلب بيانات التقرير
    report_data = get_ledger_data(conn, filters)
    conn.close()
    
    return render_template(
        'reports/ledger_report.html',
        report_data=report_data,
        filters=filters,
        default_start_date=default_start_date,
        default_end_date=default_end_date,
        parent_accounts=parent_accounts,
        account_types=ACCOUNT_TYPES
    )

@ledger_bp.route('/reports/ledger/print')
def print_ledger_report():
    """طباعة تقرير دفتر الأستاذ"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    
    if request.args.get('parent_account_id'):
        filters['parent_account_id'] = request.args.get('parent_account_id')
    
    # جلب بيانات التقرير
    conn = get_db()
    report_data = get_ledger_data(conn, filters)
    conn.close()
    
    # تاريخ الطباعة
    print_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    return render_template(
        'reports/print_ledger_report.html',
        report_data=report_data,
        filters=filters,
        print_date=print_date,
        account_types=ACCOUNT_TYPES
    )

@ledger_bp.route('/reports/ledger/export')
def export_ledger_report():
    """تصدير تقرير دفتر الأستاذ إلى Excel"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    
    if request.args.get('parent_account_id'):
        filters['parent_account_id'] = request.args.get('parent_account_id')
    
    # جلب بيانات التقرير
    conn = get_db()
    report_data = get_ledger_data(conn, filters)
    conn.close()
    
    # تحويل البيانات إلى DataFrame
    df_data = []
    
    # إضافة بيانات الحسابات
    for account in report_data['accounts']:
        df_data.append({
            'كود الحساب': account['code'],
            'اسم الحساب': account['name'],
            'رصيد أول المدة (مدين)': account['opening_debit'],
            'رصيد أول المدة (دائن)': account['opening_credit'],
            'حركة الفترة (مدين)': account['period_debit'],
            'حركة الفترة (دائن)': account['period_credit'],
            'مجموع الحركات (مدين)': account['total_debit'],
            'مجموع الحركات (دائن)': account['total_credit'],
            'الرصيد (مدين)': account['balance_debit'],
            'الرصيد (دائن)': account['balance_credit']
        })
    
    # إضافة المجاميع الكلية
    df_data.append({
        'كود الحساب': '',
        'اسم الحساب': 'الإجمالي',
        'رصيد أول المدة (مدين)': report_data['totals']['opening_debit'],
        'رصيد أول المدة (دائن)': report_data['totals']['opening_credit'],
        'حركة الفترة (مدين)': report_data['totals']['period_debit'],
        'حركة الفترة (دائن)': report_data['totals']['period_credit'],
        'مجموع الحركات (مدين)': report_data['totals']['total_debit'],
        'مجموع الحركات (دائن)': report_data['totals']['total_credit'],
        'الرصيد (مدين)': report_data['totals']['balance_debit'],
        'الرصيد (دائن)': report_data['totals']['balance_credit']
    })
    
    # إنشاء DataFrame
    df = pd.DataFrame(df_data)
    
    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()
    
    # استخدام openpyxl مع تنسيق الخلايا
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='دفتر الأستاذ', index=False)
        
        # الحصول على ورقة العمل لتطبيق التنسيق
        workbook = writer.book
        worksheet = writer.sheets['دفتر الأستاذ']
        
        # إضافة عنوان للتقرير
        # إدراج صفوف فارغة في بداية الورقة
        worksheet.insert_rows(1, 3)
        
        # إضافة عنوان التقرير
        title = "تقرير دفتر الأستاذ"
        if report_data['parent_account']:
            title += f" - {report_data['parent_account']['code']} - {report_data['parent_account']['name']}"
            
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # دمج الخلايا للعنوان
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        
        # إضافة معلومات الفترة
        date_range = "الفترة: "
        if filters.get('start_date'):
            date_range += f"من {filters['start_date']} "
        if filters.get('end_date'):
            date_range += f"إلى {filters['end_date']}"
        
        date_cell = worksheet.cell(row=2, column=1)
        date_cell.value = date_range
        date_cell.font = Font(size=12)
        date_cell.alignment = Alignment(horizontal='center')
        
        # دمج الخلايا لمعلومات الفترة
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
        
        # إضافة تاريخ الطباعة
        print_date_cell = worksheet.cell(row=3, column=1)
        print_date_cell.value = f"تاريخ الطباعة: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        print_date_cell.font = Font(size=10)
        print_date_cell.alignment = Alignment(horizontal='center')
        
        # دمج الخلايا لتاريخ الطباعة
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
        
        # تعديل رأس الجدول ليبدأ من الصف الرابع
        for cell in worksheet[4]:
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # تعيين عرض الأعمدة
        worksheet.column_dimensions['A'].width = 15  # كود الحساب
        worksheet.column_dimensions['B'].width = 30  # اسم الحساب
        worksheet.column_dimensions['C'].width = 15  # رصيد أول المدة (مدين)
        worksheet.column_dimensions['D'].width = 15  # رصيد أول المدة (دائن)
        worksheet.column_dimensions['E'].width = 15  # حركة الفترة (مدين)
        worksheet.column_dimensions['F'].width = 15  # حركة الفترة (دائن)
        worksheet.column_dimensions['G'].width = 15  # مجموع الحركات (مدين)
        worksheet.column_dimensions['H'].width = 15  # مجموع الحركات (دائن)
        worksheet.column_dimensions['I'].width = 15  # الرصيد (مدين)
        worksheet.column_dimensions['J'].width = 15  # الرصيد (دائن)
        
        # تنسيق الخلايا
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تعريف الألوان والأنماط
        total_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        
        # تنسيق باقي الخلايا (بدءًا من الصف الخامس بعد إضافة العنوان)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=5, max_row=worksheet.max_row), 5):
            for cell in row:
                cell.border = border
                
                # تنسيق الأرقام
                if cell.column > 2:  # الأعمدة المالية
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.alignment = Alignment(horizontal='right')
            
            # تنسيق صف المجموع
            if row_idx == worksheet.max_row:
                for cell in row:
                    cell.fill = total_fill
                    cell.font = Font(bold=True)
    
    # إعادة مؤشر الملف إلى البداية
    output.seek(0)
    
    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.now()
    file_name = f"دفتر_الأستاذ_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # إرسال الملف للتنزيل
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@ledger_bp.route('/reports/ledger/search_parent_accounts', methods=['GET'])
def search_parent_accounts():
    """البحث عن الحسابات الرئيسية بناءً على معايير البحث"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return {'success': False, 'message': 'يرجى تسجيل الدخول أولاً'}, 401
    
    # استخراج معايير البحث
    search_term = request.args.get('search_term', '')
    
    # الاتصال بقاعدة البيانات
    conn = get_db()
    cursor = conn.cursor()
    
    # البحث عن الحسابات الرئيسية التي تطابق معايير البحث
    cursor.execute("""
        SELECT DISTINCT a.id, a.code, a.name, a.account_type
        FROM accounts a
        WHERE EXISTS (
            SELECT 1 FROM accounts sub WHERE sub.parent_id = a.id
        )
        AND a.is_active = 1
        AND (a.code LIKE ? OR a.name LIKE ?)
        ORDER BY a.code
        LIMIT 50
    """, (f'%{search_term}%', f'%{search_term}%'))
    
    # استخراج النتائج
    results = cursor.fetchall()
    
    # تحويل النتائج إلى قائمة من القواميس
    accounts = []
    for row in results:
        accounts.append({
            'id': row['id'],
            'code': row['code'],
            'name': row['name'],
            'account_type': row['account_type'],
            'account_type_name': ACCOUNT_TYPES.get(row['account_type'], '')
        })
    
    # إغلاق الاتصال بقاعدة البيانات
    conn.close()
    
    # إرجاع النتائج كـ JSON
    return {'success': True, 'accounts': accounts}
