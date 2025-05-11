# e:\0\7-5\V9\profit_loss_by_cost_center.py
from flask import Blueprint, render_template, request, redirect, url_for, session, send_file
from database import get_db, ACCOUNT_TYPES, ACCOUNT_TYPES_REVERSE
import pandas as pd
import io
from datetime import datetime

# إنشاء Blueprint لتقرير الأرباح والخسائر حسب مراكز التكلفة
profit_loss_by_cc_bp = Blueprint('profit_loss_by_cc', __name__)

def get_active_financial_year_dates():
    """
    الحصول على تواريخ بداية ونهاية السنة المالية النشطة
    
    Returns:
    - tuple: (start_date, end_date) أو (None, None) إذا لم يتم العثور على سنة مالية نشطة
    """
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT start_date, end_date 
            FROM financial_years 
            WHERE is_active = 1 
            LIMIT 1
        """)
        
        result = cursor.fetchone()
        if result:
            return result['start_date'], result['end_date']
        return None, None
    finally:
        conn.close()

def get_cost_centers(conn):
    """
    استخراج قائمة مراكز التكلفة
    
    Parameters:
    - conn: اتصال قاعدة البيانات
    
    Returns:
    - قائمة بمراكز التكلفة
    """
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, code FROM cost_centers ORDER BY name")
    return cursor.fetchall()

def get_profit_loss_data_by_cost_center(conn, filters, cost_center_id=None):
    """
    استخراج بيانات تقرير الأرباح والخسائر لمركز تكلفة محدد
    
    Parameters:
    - conn: اتصال قاعدة البيانات
    - filters: معايير التصفية (start_date, end_date)
    - cost_center_id: معرف مركز التكلفة (إذا كان None، يتم استخراج البيانات بدون تصفية حسب مركز التكلفة)
    
    Returns:
    - قاموس يحتوي على بيانات الإيرادات والمصروفات والأرباح/الخسائر
    """
    cursor = conn.cursor()
    
    # استخراج الإيرادات (نوع الحساب = 4)
    revenue_type = ACCOUNT_TYPES_REVERSE['إيرادات']
    
    # استعلام الإيرادات
    revenue_query = """
    SELECT 
        a.id, a.code, a.name,
        SUM(CASE WHEN jvd.credit > 0 THEN jvd.credit ELSE 0 END) - 
        SUM(CASE WHEN jvd.debit > 0 THEN jvd.debit ELSE 0 END) as amount
    FROM 
        accounts a
    LEFT JOIN 
        journal_voucher_details jvd ON a.id = jvd.account_id
    LEFT JOIN 
        journal_vouchers jv ON jvd.voucher_id = jv.id
    WHERE 
        a.account_type = ? AND a.is_active = 1
    """
    
    params = [revenue_type]
    
    # إضافة شرط مركز التكلفة إذا تم تحديده
    if cost_center_id:
        revenue_query += " AND jvd.cost_center_id = ?"
        params.append(cost_center_id)
    
    # إضافة شروط التصفية إذا وجدت
    if filters.get('start_date'):
        revenue_query += " AND jv.voucher_date >= ?"
        params.append(filters['start_date'])
    
    if filters.get('end_date'):
        revenue_query += " AND jv.voucher_date <= ?"
        params.append(filters['end_date'])
    
    revenue_query += " GROUP BY a.id, a.code, a.name ORDER BY a.code"
    
    cursor.execute(revenue_query, params)
    revenues = cursor.fetchall()
    
    # استخراج المصروفات (نوع الحساب = 5)
    expense_type = ACCOUNT_TYPES_REVERSE['مصروفات']
    
    # استعلام المصروفات
    expense_query = """
    SELECT 
        a.id, a.code, a.name,
        SUM(CASE WHEN jvd.debit > 0 THEN jvd.debit ELSE 0 END) - 
        SUM(CASE WHEN jvd.credit > 0 THEN jvd.credit ELSE 0 END) as amount
    FROM 
        accounts a
    LEFT JOIN 
        journal_voucher_details jvd ON a.id = jvd.account_id
    LEFT JOIN 
        journal_vouchers jv ON jvd.voucher_id = jv.id
    WHERE 
        a.account_type = ? AND a.is_active = 1
    """
    
    params = [expense_type]
    
    # إضافة شرط مركز التكلفة إذا تم تحديده
    if cost_center_id:
        expense_query += " AND jvd.cost_center_id = ?"
        params.append(cost_center_id)
    
    # إضافة شروط التصفية إذا وجدت
    if filters.get('start_date'):
        expense_query += " AND jv.voucher_date >= ?"
        params.append(filters['start_date'])
    
    if filters.get('end_date'):
        expense_query += " AND jv.voucher_date <= ?"
        params.append(filters['end_date'])
    
    expense_query += " GROUP BY a.id, a.code, a.name ORDER BY a.code"
    
    cursor.execute(expense_query, params)
    expenses = cursor.fetchall()
    
    # حساب إجمالي الإيرادات
    total_revenue = sum(rev['amount'] for rev in revenues)
    
    # حساب إجمالي المصروفات
    total_expense = sum(exp['amount'] for exp in expenses)
    
    # حساب صافي الربح أو الخسارة
    net_profit = total_revenue - total_expense
    
    return {
        'revenues': revenues,
        'expenses': expenses,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'is_profit': net_profit >= 0  # علامة لتحديد ما إذا كان ربح أو خسارة
    }

@profit_loss_by_cc_bp.route('/reports/profit_loss_by_cost_center')
def profit_loss_by_cost_center_view():
    """عرض صفحة تقرير الأرباح والخسائر حسب مراكز التكلفة مع إمكانية التصفية"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # إضافة معيار تصفية مركز التكلفة إذا تم تحديده
    if request.args.get('cost_center_id'):
        filters['cost_center_id'] = request.args.get('cost_center_id')
    
    # جلب بيانات التقرير
    conn = get_db()
    
    # الحصول على قائمة مراكز التكلفة
    cost_centers = get_cost_centers(conn)
    
    # إنشاء قاموس لتخزين بيانات التقرير لكل مركز تكلفة
    reports_by_cost_center = []
    
    # إذا تم تحديد مركز تكلفة معين، نعرض فقط تقرير ذلك المركز
    if filters.get('cost_center_id'):
        # البحث عن مركز التكلفة المحدد
        selected_cost_center = None
        for cc in cost_centers:
            if str(cc['id']) == str(filters['cost_center_id']):
                selected_cost_center = cc
                break
        
        if selected_cost_center:
            # إضافة تقرير لمركز التكلفة المحدد
            report_data = get_profit_loss_data_by_cost_center(conn, filters, selected_cost_center['id'])
            reports_by_cost_center.append({
                'cost_center': selected_cost_center,
                'report_data': report_data
            })
    else:
        # إضافة تقرير إجمالي (بدون تصفية حسب مركز التكلفة)
        total_report = get_profit_loss_data_by_cost_center(conn, filters)
        reports_by_cost_center.append({
            'cost_center': {'id': None, 'name': 'الإجمالي (كل مراكز التكلفة)', 'code': ''},
            'report_data': total_report
        })
        
        # إضافة تقرير لكل مركز تكلفة
        for cost_center in cost_centers:
            report_data = get_profit_loss_data_by_cost_center(conn, filters, cost_center['id'])
            reports_by_cost_center.append({
                'cost_center': cost_center,
                'report_data': report_data
            })
    
    conn.close()
    
    return render_template(
        'reports/profit_loss_by_cost_center_report.html',
        reports_by_cost_center=reports_by_cost_center,
        filters=filters,
        default_start_date=default_start_date,
        default_end_date=default_end_date,
        all_cost_centers=cost_centers
    )

@profit_loss_by_cc_bp.route('/reports/profit_loss_by_cost_center/print')
def print_profit_loss_by_cost_center_report():
    """عرض نسخة للطباعة من تقرير الأرباح والخسائر حسب مراكز التكلفة"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # إضافة معيار تصفية مركز التكلفة إذا تم تحديده
    if request.args.get('cost_center_id'):
        filters['cost_center_id'] = request.args.get('cost_center_id')
    
    # جلب بيانات التقرير
    conn = get_db()
    
    # الحصول على قائمة مراكز التكلفة
    cost_centers = get_cost_centers(conn)
    
    # إنشاء قاموس لتخزين بيانات التقرير لكل مركز تكلفة
    reports_by_cost_center = []
    
    # إذا تم تحديد مركز تكلفة معين، نعرض فقط تقرير ذلك المركز
    if filters.get('cost_center_id'):
        # البحث عن مركز التكلفة المحدد
        selected_cost_center = None
        for cc in cost_centers:
            if str(cc['id']) == str(filters['cost_center_id']):
                selected_cost_center = cc
                break
        
        if selected_cost_center:
            # إضافة تقرير لمركز التكلفة المحدد
            report_data = get_profit_loss_data_by_cost_center(conn, filters, selected_cost_center['id'])
            reports_by_cost_center.append({
                'cost_center': selected_cost_center,
                'report_data': report_data
            })
    else:
        # إضافة تقرير إجمالي (بدون تصفية حسب مركز التكلفة)
        total_report = get_profit_loss_data_by_cost_center(conn, filters)
        reports_by_cost_center.append({
            'cost_center': {'id': None, 'name': 'الإجمالي (كل مراكز التكلفة)', 'code': ''},
            'report_data': total_report
        })
        
        # إضافة تقرير لكل مركز تكلفة
        for cost_center in cost_centers:
            report_data = get_profit_loss_data_by_cost_center(conn, filters, cost_center['id'])
            reports_by_cost_center.append({
                'cost_center': cost_center,
                'report_data': report_data
            })
    
    conn.close()
    
    # تاريخ الطباعة
    print_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    return render_template(
        'reports/print_profit_loss_by_cost_center_report.html',
        reports_by_cost_center=reports_by_cost_center,
        filters=filters,
        print_date=print_date
    )

@profit_loss_by_cc_bp.route('/reports/profit_loss_by_cost_center/export')
def export_profit_loss_by_cost_center_report():
    """تصدير تقرير الأرباح والخسائر حسب مراكز التكلفة إلى Excel"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # إضافة معيار تصفية مركز التكلفة إذا تم تحديده
    if request.args.get('cost_center_id'):
        filters['cost_center_id'] = request.args.get('cost_center_id')
    
    # جلب بيانات التقرير
    conn = get_db()
    
    # الحصول على قائمة مراكز التكلفة
    cost_centers = get_cost_centers(conn)
    
    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()
    
    # استخدام openpyxl مع تنسيق الخلايا
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # إنشاء ورقة للملخص الشامل
        summary_sheet_name = 'ملخص شامل'
        
        # إنشاء قائمة لتخزين بيانات جميع مراكز التكلفة
        all_centers_data = []
        
        # إضافة عنوان التقرير
        all_centers_data.append({
            'مركز التكلفة': 'تقرير الأرباح والخسائر حسب مراكز التكلفة',
            'البند': '',
            'المبلغ': None
        })
        
        # إضافة تاريخ التقرير
        current_date = datetime.now().strftime('%Y-%m-%d')
        from_date = filters.get('start_date', 'غير محدد')
        to_date = filters.get('end_date', 'غير محدد')
        
        all_centers_data.append({
            'مركز التكلفة': f'الفترة من: {from_date} إلى: {to_date}',
            'البند': '',
            'المبلغ': None
        })
        
        all_centers_data.append({
            'مركز التكلفة': f'تاريخ إنشاء التقرير: {current_date}',
            'البند': '',
            'المبلغ': None
        })
        
        # إضافة سطر فارغ
        all_centers_data.append({
            'مركز التكلفة': '',
            'البند': '',
            'المبلغ': None
        })
        
        # إذا تم تحديد مركز تكلفة معين، نصدر فقط تقرير ذلك المركز
        if filters.get('cost_center_id'):
            # البحث عن مركز التكلفة المحدد
            selected_cost_center = None
            for cc in cost_centers:
                if str(cc['id']) == str(filters['cost_center_id']):
                    selected_cost_center = cc
                    break
            
            if selected_cost_center:
                # إضافة تقرير لمركز التكلفة المحدد
                report_data = get_profit_loss_data_by_cost_center(conn, filters, selected_cost_center['id'])
                
                # إضافة بيانات مركز التكلفة إلى الملخص الشامل
                cc_name = selected_cost_center['name']
                if selected_cost_center['code']:
                    cc_name = f"{selected_cost_center['code']} - {cc_name}"
                
                # إضافة عنوان مركز التكلفة
                all_centers_data.append({
                    'مركز التكلفة': cc_name,
                    'البند': '',
                    'المبلغ': None
                })
                
                # إضافة الإيرادات
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': 'اجمالي الإيرادات',
                    'المبلغ': report_data['total_revenue']
                })
                
                # إضافة تفاصيل الإيرادات
                for rev in report_data['revenues']:
                    all_centers_data.append({
                        'مركز التكلفة': '',
                        'البند': '    ' + rev['name'],
                        'المبلغ': rev['amount']
                    })
                
                # إضافة المصروفات
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': 'اجمالي المصروفات',
                    'المبلغ': report_data['total_expense']
                })
                
                # إضافة تفاصيل المصروفات
                for exp in report_data['expenses']:
                    all_centers_data.append({
                        'مركز التكلفة': '',
                        'البند': '    ' + exp['name'],
                        'المبلغ': exp['amount']
                    })
                
                # إضافة صافي الربح أو الخسارة
                profit_label = 'الأرباح' if report_data['is_profit'] else 'الخسائر'
                profit_value = report_data['net_profit'] if report_data['is_profit'] else -report_data['net_profit']
                
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': profit_label,
                    'المبلغ': profit_value
                })
                
                # إضافة سطر فارغ
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': '',
                    'المبلغ': None
                })
                
                # إضافة ورقة تفصيلية لمركز التكلفة
                sheet_name = selected_cost_center['name']
                if len(sheet_name) > 31:  # Excel يحد اسم ورقة العمل بـ 31 حرف
                    sheet_name = sheet_name[:28] + '...'
                export_profit_loss_to_excel(writer, sheet_name, report_data)
        else:
            # أولاً: إضافة بيانات الإجمالي إلى الملخص الشامل
            total_report = get_profit_loss_data_by_cost_center(conn, filters)
            
            # إضافة عنوان الإجمالي
            all_centers_data.append({
                'مركز التكلفة': 'الإجمالي (كل مراكز التكلفة)',
                'البند': '',
                'المبلغ': None
            })
            
            # إضافة الإيرادات
            all_centers_data.append({
                'مركز التكلفة': '',
                'البند': 'اجمالي الإيرادات',
                'المبلغ': total_report['total_revenue']
            })
            
            # إضافة تفاصيل الإيرادات
            for rev in total_report['revenues']:
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': '    ' + rev['name'],
                    'المبلغ': rev['amount']
                })
            
            # إضافة المصروفات
            all_centers_data.append({
                'مركز التكلفة': '',
                'البند': 'اجمالي المصروفات',
                'المبلغ': total_report['total_expense']
            })
            
            # إضافة تفاصيل المصروفات
            for exp in total_report['expenses']:
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': '    ' + exp['name'],
                    'المبلغ': exp['amount']
                })
            
            # إضافة صافي الربح أو الخسارة
            profit_label = 'الأرباح' if total_report['is_profit'] else 'الخسائر'
            profit_value = total_report['net_profit'] if total_report['is_profit'] else -total_report['net_profit']
            
            all_centers_data.append({
                'مركز التكلفة': '',
                'البند': profit_label,
                'المبلغ': profit_value
            })
            
            # إضافة سطر فارغ
            all_centers_data.append({
                'مركز التكلفة': '',
                'البند': '',
                'المبلغ': None
            })
            
            # إضافة ورقة تفصيلية للإجمالي
            export_profit_loss_to_excel(writer, 'الإجمالي', total_report)
            
            # ثانياً: إضافة بيانات كل مركز تكلفة إلى الملخص الشامل وإنشاء ورقة تفصيلية لكل مركز
            for cost_center in cost_centers:
                report_data = get_profit_loss_data_by_cost_center(conn, filters, cost_center['id'])
                
                # إضافة بيانات مركز التكلفة إلى الملخص الشامل
                cc_name = cost_center['name']
                if cost_center['code']:
                    cc_name = f"{cost_center['code']} - {cc_name}"
                
                # إضافة عنوان مركز التكلفة
                all_centers_data.append({
                    'مركز التكلفة': cc_name,
                    'البند': '',
                    'المبلغ': None
                })
                
                # إضافة الإيرادات
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': 'اجمالي الإيرادات',
                    'المبلغ': report_data['total_revenue']
                })
                
                # إضافة تفاصيل الإيرادات
                for rev in report_data['revenues']:
                    all_centers_data.append({
                        'مركز التكلفة': '',
                        'البند': '    ' + rev['name'],
                        'المبلغ': rev['amount']
                    })
                
                # إضافة المصروفات
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': 'اجمالي المصروفات',
                    'المبلغ': report_data['total_expense']
                })
                
                # إضافة تفاصيل المصروفات
                for exp in report_data['expenses']:
                    all_centers_data.append({
                        'مركز التكلفة': '',
                        'البند': '    ' + exp['name'],
                        'المبلغ': exp['amount']
                    })
                
                # إضافة صافي الربح أو الخسارة
                profit_label = 'الأرباح' if report_data['is_profit'] else 'الخسائر'
                profit_value = report_data['net_profit'] if report_data['is_profit'] else -report_data['net_profit']
                
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': profit_label,
                    'المبلغ': profit_value
                })
                
                # إضافة سطر فارغ
                all_centers_data.append({
                    'مركز التكلفة': '',
                    'البند': '',
                    'المبلغ': None
                })
                
                # إضافة ورقة تفصيلية لمركز التكلفة
                sheet_name = cost_center['name']
                if len(sheet_name) > 31:  # Excel يحد اسم ورقة العمل بـ 31 حرف
                    sheet_name = sheet_name[:28] + '...'
                export_profit_loss_to_excel(writer, sheet_name, report_data)
        
        # إنشاء DataFrame للملخص الشامل وكتابته إلى ورقة العمل
        summary_df = pd.DataFrame(all_centers_data)
        summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
        
        # تنسيق ورقة الملخص الشامل
        worksheet = writer.sheets[summary_sheet_name]
        
        # تعيين عرض الأعمدة
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 20
        
        # تطبيق التنسيق على الخلايا
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # تعريف الألوان والأنماط
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        
        # تنسيق العنوان
        for row in range(1, 4):
            cell = worksheet.cell(row=row, column=1)
            cell.font = title_font
            for col in range(1, 4):
                worksheet.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        # تنسيق بيانات الملخص الشامل
        for row in range(1, len(all_centers_data) + 1):
            for col in range(1, 4):
                cell = worksheet.cell(row=row, column=col)
                
                # تنسيق عناوين مراكز التكلفة
                if col == 1 and cell.value and cell.value != '':
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                
                # تنسيق عناوين الإيرادات والمصروفات
                if col == 2 and cell.value in ['اجمالي الإيرادات', 'اجمالي المصروفات', 'الأرباح', 'الخسائر']:
                    cell.font = Font(bold=True)
                    
                    if cell.value == 'اجمالي الإيرادات':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif cell.value == 'اجمالي المصروفات':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif cell.value == 'الأرباح':
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif cell.value == 'الخسائر':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
                # محاذاة النص إلى اليمين والأرقام إلى اليسار
                if col in [1, 2]:
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
                
                # إضافة حدود للخلايا
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    conn.close()
    
    # إعادة مؤشر الملف إلى البداية
    output.seek(0)
    
    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.now()
    file_name = f"تقرير_الأرباح_والخسائر_حسب_مراكز_التكلفة_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # إرسال الملف للتنزيل
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_profit_loss_to_excel(writer, sheet_name, report_data):
    """
    تصدير بيانات تقرير الأرباح والخسائر إلى ورقة عمل Excel
    
    Parameters:
    - writer: كائن ExcelWriter
    - sheet_name: اسم ورقة العمل
    - report_data: بيانات التقرير
    """
    # إنشاء DataFrame واحد يحتوي على جميع البيانات
    all_data = []
    
    # إضافة عنوان التقرير
    all_data.append({
        'البند': 'تقرير الأرباح والخسائر حسب مراكز التكلفة',
        'المبلغ': None
    })
    
    # إضافة تاريخ التقرير
    current_date = datetime.now().strftime('%Y-%m-%d')
    all_data.append({
        'البند': f'تاريخ التقرير: {current_date}',
        'المبلغ': None
    })
    
    # إضافة سطر فارغ
    all_data.append({
        'البند': '',
        'المبلغ': None
    })
    
    # إضافة عنوان قسم الإيرادات
    all_data.append({
        'البند': 'اجمالي الإيرادات',
        'المبلغ': report_data['total_revenue']
    })
    
    # إضافة تفاصيل الإيرادات
    for rev in report_data['revenues']:
        all_data.append({
            'البند': '    ' + rev['name'],  # إضافة مسافات للتمييز
            'المبلغ': rev['amount']
        })
    
    # إضافة سطر فارغ
    all_data.append({
        'البند': '',
        'المبلغ': None
    })
    
    # إضافة عنوان قسم المصروفات
    all_data.append({
        'البند': 'اجمالي المصروفات',
        'المبلغ': report_data['total_expense']
    })
    
    # إضافة تفاصيل المصروفات
    for exp in report_data['expenses']:
        all_data.append({
            'البند': '    ' + exp['name'],  # إضافة مسافات للتمييز
            'المبلغ': exp['amount']
        })
    
    # إضافة سطر فارغ
    all_data.append({
        'البند': '',
        'المبلغ': None
    })
    
    # إضافة صافي الربح أو الخسارة
    profit_label = 'الأرباح' if report_data['is_profit'] else 'الخسائر'
    profit_value = report_data['net_profit'] if report_data['is_profit'] else -report_data['net_profit']
    
    all_data.append({
        'البند': profit_label,
        'المبلغ': profit_value
    })
    
    # تحويل البيانات إلى DataFrame
    df = pd.DataFrame(all_data)
    
    # كتابة البيانات إلى ورقة العمل
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # الحصول على ورقة العمل لتطبيق التنسيق
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # تعيين عرض الأعمدة
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 20
    
    # تطبيق التنسيق على الخلايا
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # تعريف الألوان والأنماط
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    revenue_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    expense_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    profit_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    loss_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # تنسيق الخلايا
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # تطبيق التنسيق على الخلايا
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
        # تطبيق المحاذاة والحدوب على جميع الخلايا
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            # تنسيق الأرقام
            if cell.column == 2 and cell.value is not None:  # عمود المبلغ
                cell.number_format = '#,##0.00'
        
        # تطبيق التنسيق حسب نوع البند
        cell_value = worksheet.cell(row=row_idx, column=1).value
        amount_cell = worksheet.cell(row=row_idx, column=2)
        
        if cell_value == 'اجمالي الإيرادات':
            worksheet.cell(row=row_idx, column=1).fill = revenue_fill
            worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
            amount_cell.fill = revenue_fill
            amount_cell.font = Font(bold=True)
        elif cell_value == 'اجمالي المصروفات':
            worksheet.cell(row=row_idx, column=1).fill = expense_fill
            worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
            amount_cell.fill = expense_fill
            amount_cell.font = Font(bold=True)
        elif cell_value == 'الأرباح':
            worksheet.cell(row=row_idx, column=1).fill = profit_fill
            worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
            amount_cell.fill = profit_fill
            amount_cell.font = Font(bold=True)
        elif cell_value == 'الخسائر':
            worksheet.cell(row=row_idx, column=1).fill = loss_fill
            worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
            amount_cell.fill = loss_fill
            amount_cell.font = Font(bold=True)