# e:\0\6-5\V9\profit_loss_report.py
from flask import Blueprint, render_template, request, redirect, url_for, session, send_file
from database import get_db, ACCOUNT_TYPES, ACCOUNT_TYPES_REVERSE
import pandas as pd
import io
from datetime import datetime

# إنشاء Blueprint لتقرير الأرباح والخسائر
profit_loss_bp = Blueprint('profit_loss', __name__)

def get_active_financial_year_dates():
    """
    الحصول على تواريخ بداية ونهاية السنة المالية النشطة
    
    Returns:
    - tuple: (start_date, end_date) أو (None, None) إذا لم يتم العثور على سنة مالية نشطة
    """
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT start_date, end_date 
            FROM financial_years 
            WHERE is_active = 1 
            LIMIT 1
        """)
        
        result = cursor.fetchone()
        if result:
            return result['start_date'], result['end_date']
        return None, None
    finally:
        conn.close()

def get_profit_loss_data(conn, filters):
    """
    استخراج بيانات تقرير الأرباح والخسائر
    
    Parameters:
    - conn: اتصال قاعدة البيانات
    - filters: معايير التصفية (start_date, end_date)
    
    Returns:
    - قاموس يحتوي على بيانات الإيرادات والمصروفات والأرباح/الخسائر
    """
    cursor = conn.cursor()
    
    # استخراج الإيرادات (نوع الحساب = 4)
    revenue_type = ACCOUNT_TYPES_REVERSE['إيرادات']
    
    # استعلام الإيرادات
    revenue_query = """
    SELECT 
        a.id, a.code, a.name,
        SUM(CASE WHEN jvd.credit > 0 THEN jvd.credit ELSE 0 END) - 
        SUM(CASE WHEN jvd.debit > 0 THEN jvd.debit ELSE 0 END) as amount
    FROM 
        accounts a
    LEFT JOIN 
        journal_voucher_details jvd ON a.id = jvd.account_id
    LEFT JOIN 
        journal_vouchers jv ON jvd.voucher_id = jv.id
    WHERE 
        a.account_type = ? AND a.is_active = 1
    """
    
    params = [revenue_type]
    
    # إضافة شروط التصفية إذا وجدت
    if filters.get('start_date'):
        revenue_query += " AND jv.voucher_date >= ?"
        params.append(filters['start_date'])
    
    if filters.get('end_date'):
        revenue_query += " AND jv.voucher_date <= ?"
        params.append(filters['end_date'])
    
    revenue_query += " GROUP BY a.id, a.code, a.name ORDER BY a.code"
    
    cursor.execute(revenue_query, params)
    revenues = cursor.fetchall()
    
    # استخراج المصروفات (نوع الحساب = 5)
    expense_type = ACCOUNT_TYPES_REVERSE['مصروفات']
    
    # استعلام المصروفات
    expense_query = """
    SELECT 
        a.id, a.code, a.name,
        SUM(CASE WHEN jvd.debit > 0 THEN jvd.debit ELSE 0 END) - 
        SUM(CASE WHEN jvd.credit > 0 THEN jvd.credit ELSE 0 END) as amount
    FROM 
        accounts a
    LEFT JOIN 
        journal_voucher_details jvd ON a.id = jvd.account_id
    LEFT JOIN 
        journal_vouchers jv ON jvd.voucher_id = jv.id
    WHERE 
        a.account_type = ? AND a.is_active = 1
    """
    
    params = [expense_type]
    
    # إضافة شروط التصفية إذا وجدت
    if filters.get('start_date'):
        expense_query += " AND jv.voucher_date >= ?"
        params.append(filters['start_date'])
    
    if filters.get('end_date'):
        expense_query += " AND jv.voucher_date <= ?"
        params.append(filters['end_date'])
    
    expense_query += " GROUP BY a.id, a.code, a.name ORDER BY a.code"
    
    cursor.execute(expense_query, params)
    expenses = cursor.fetchall()
    
    # حساب إجمالي الإيرادات
    total_revenue = sum(rev['amount'] for rev in revenues)
    
    # حساب إجمالي المصروفات
    total_expense = sum(exp['amount'] for exp in expenses)
    
    # حساب صافي الربح أو الخسارة
    net_profit = total_revenue - total_expense
    
    return {
        'revenues': revenues,
        'expenses': expenses,
        'total_revenue': total_revenue,
        'total_expense': total_expense,
        'net_profit': net_profit,
        'is_profit': net_profit >= 0  # علامة لتحديد ما إذا كان ربح أو خسارة
    }

@profit_loss_bp.route('/reports/profit_loss')
def profit_loss_report_view():
    """عرض صفحة تقرير الأرباح والخسائر مع إمكانية التصفية"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # جلب بيانات التقرير
    conn = get_db()
    report_data = get_profit_loss_data(conn, filters)
    conn.close()
    
    return render_template(
        'reports/profit_loss_report.html',
        report_data=report_data,
        filters=filters,
        default_start_date=default_start_date,
        default_end_date=default_end_date
    )

@profit_loss_bp.route('/reports/profit_loss/print')
def print_profit_loss_report():
    """عرض نسخة للطباعة من تقرير الأرباح والخسائر"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # جلب بيانات التقرير
    conn = get_db()
    report_data = get_profit_loss_data(conn, filters)
    conn.close()
    
    # تاريخ الطباعة
    print_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    return render_template(
        'reports/print_profit_loss_report.html',
        report_data=report_data,
        filters=filters,
        print_date=print_date
    )

@profit_loss_bp.route('/reports/profit_loss/export')
def export_profit_loss_report():
    """تصدير تقرير الأرباح والخسائر إلى Excel"""
    # التحقق من تسجيل الدخول
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # الحصول على تواريخ السنة المالية النشطة
    default_start_date, default_end_date = get_active_financial_year_dates()
    
    # استخراج معايير التصفية من الطلب
    filters = {}
    
    # استخدام تاريخ بداية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ البداية
    if request.args.get('start_date'):
        filters['start_date'] = request.args.get('start_date')
    elif default_start_date:
        filters['start_date'] = default_start_date
        
    # استخدام تاريخ نهاية السنة المالية النشطة كقيمة افتراضية إذا لم يتم تحديد تاريخ النهاية
    if request.args.get('end_date'):
        filters['end_date'] = request.args.get('end_date')
    elif default_end_date:
        filters['end_date'] = default_end_date
    
    # جلب بيانات التقرير
    conn = get_db()
    report_data = get_profit_loss_data(conn, filters)
    conn.close()
    
    # إنشاء DataFrame واحد يحتوي على جميع البيانات
    # إضافة عنوان للإيرادات
    all_data = []
    
    # إضافة عنوان قسم الإيرادات
    all_data.append({
        'البند': 'اجمالي الإيرادات',
        'المبلغ': report_data['total_revenue']
    })
    
    # إضافة تفاصيل الإيرادات
    for rev in report_data['revenues']:
        all_data.append({
            'البند': '    ' + rev['name'],  # إضافة مسافات للتمييز
            'المبلغ': rev['amount']
        })
    
    # إضافة سطر فارغ
    all_data.append({
        'البند': '',
        'المبلغ': None
    })
    
    # إضافة عنوان قسم المصروفات
    all_data.append({
        'البند': 'اجمالي المصروفات',
        'المبلغ': report_data['total_expense']
    })
    
    # إضافة تفاصيل المصروفات
    for exp in report_data['expenses']:
        all_data.append({
            'البند': '    ' + exp['name'],  # إضافة مسافات للتمييز
            'المبلغ': exp['amount']
        })
    
    # إضافة سطر فارغ
    all_data.append({
        'البند': '',
        'المبلغ': None
    })
    
    # إضافة صافي الربح أو الخسارة
    profit_label = 'الأرباح' if report_data['is_profit'] else 'الخسائر'
    profit_value = report_data['net_profit'] if report_data['is_profit'] else -report_data['net_profit']
    
    all_data.append({
        'البند': profit_label,
        'المبلغ': profit_value
    })
    
    # تحويل البيانات إلى DataFrame
    df = pd.DataFrame(all_data)
    
    # إنشاء ملف Excel في الذاكرة
    output = io.BytesIO()
    
    # استخدام openpyxl مع تنسيق الخلايا
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='تقرير الأرباح والخسائر', index=False)
        
        # الحصول على ورقة العمل لتطبيق التنسيق
        workbook = writer.book
        worksheet = writer.sheets['تقرير الأرباح والخسائر']
        
        # تعيين عرض الأعمدة
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 20
        
        # تطبيق التنسيق على الخلايا
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # تعريف الألوان والأنماط
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        revenue_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        expense_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        profit_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        loss_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # تنسيق الخلايا
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # تطبيق التنسيق على الخلايا
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
            # تطبيق المحاذاة والحدوب على جميع الخلايا
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
                
                # تنسيق الأرقام
                if cell.column == 2 and cell.value is not None:  # عمود المبلغ
                    cell.number_format = '#,##0.00'
            
            # تطبيق التنسيق حسب نوع البند
            cell_value = worksheet.cell(row=row_idx, column=1).value
            amount_cell = worksheet.cell(row=row_idx, column=2)
            
            if cell_value == 'اجمالي الإيرادات':
                worksheet.cell(row=row_idx, column=1).fill = revenue_fill
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
                amount_cell.fill = revenue_fill
                amount_cell.font = Font(bold=True)
            elif cell_value == 'اجمالي المصروفات':
                worksheet.cell(row=row_idx, column=1).fill = expense_fill
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
                amount_cell.fill = expense_fill
                amount_cell.font = Font(bold=True)
            elif cell_value == 'الأرباح':
                worksheet.cell(row=row_idx, column=1).fill = profit_fill
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
                amount_cell.fill = profit_fill
                amount_cell.font = Font(bold=True)
            elif cell_value == 'الخسائر':
                worksheet.cell(row=row_idx, column=1).fill = loss_fill
                worksheet.cell(row=row_idx, column=1).font = Font(bold=True)
                amount_cell.fill = loss_fill
                amount_cell.font = Font(bold=True)
    
    # إعادة مؤشر الملف إلى البداية
    output.seek(0)
    
    # إنشاء اسم الملف مع التاريخ والوقت
    now = datetime.now()
    file_name = f"تقرير_الأرباح_والخسائر_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # إرسال الملف للتنزيل
    return send_file(
        output,
        as_attachment=True,
        download_name=file_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )